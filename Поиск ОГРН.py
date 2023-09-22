from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

# Путь к файлу Excel с номерами ИНН
excel_file_path = r'C:\Users\Grechany\Documents\python\ИНН.xlsx'

# Имя листа в Excel, на котором находятся номера ИНН
sheet_name = 'Лист1'

# Создаем экземпляр браузера Chrome
driver = webdriver.Chrome()

# Переходим на сайт
driver.get("https://egrul.nalog.ru/index.html")

# Чтение номеров ИНН из Excel
wb = openpyxl.load_workbook(excel_file_path)
sheet = wb[sheet_name]

# Создаем новую книгу Excel для результатов
result_wb = openpyxl.Workbook()
result_sheet = result_wb.active

# Начинаем с первой строки
current_row = 1

# Проходимся по каждой строке в таблице Excel
for row in sheet.iter_rows(values_only=True):
    inn = str(row[0])  # Предполагается, что номер ИНН находится в первой колонке
    search_input = driver.find_element(By.ID, "query")
    search_input.clear()  # Очищаем поле ввода
    search_input.send_keys(inn)  # Вводим ИНН

    # Добавляем задержку в 1 секунду перед нажатием кнопки "НАЙТИ"
    time.sleep(1)

    # Находим кнопку "НАЙТИ" по ID и кликаем на нее
    search_button = driver.find_element(By.ID, "btnSearch")
    search_button.click()

    try:
        # Ждем, пока элемент с id "resultPanel" станет видимым
        result_panel = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.ID, "resultPanel")))

        # Ищем контейнеры с результатами (div с class "res-row")
        result_containers = driver.find_elements(By.CLASS_NAME, "res-row")

        result_text = ""
        if result_containers:
            # Извлекаем только первый контейнер с результатом
            first_result_container = result_containers[0]

            # Получаем текст содержимого первого результата
            result_text = first_result_container.text.replace('\n', ' ')
        else:
            # Если список результатов пуст, записываем сообщение об этом
            result_text = "Данных не найдено"

        # Выводим текст результата в консоль для отладки или обработки
        print(result_text)

        # Добавляем результат в таблицу "Результат"
        result_sheet.append([result_text])

    except Exception as e:
        # Если возникла какая-либо ошибка, выводим сообщение об ошибке
        print(f"Произошла ошибка: {str(e)}")

    # Добавляем задержку в 5 секунд перед перезагрузкой страницы
    time.sleep(5)

    # Перезагружаем страницу
    driver.refresh()

# Сохраняем результаты в файл "Результат.xlsx"
result_wb.save('Результат.xlsx')

# Закрываем браузер
driver.quit()
